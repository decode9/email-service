import smtplib
from email import message
from .utils import get_template
from openpyxl import load_workbook

def init():
    email_content = get_template('promo_template')
    s = smtplib.SMTP('smtp.gmail.com: 587')
    s.starttls()
    password = "Miami2021"

    # Login Credentials for sending the mail
    s.login('pocketvzla@gmail.com', password)

    workbook = load_workbook(filename="customers_export_1.xlsx")
    sheet = workbook.active
    row_sheet = sheet.iter_rows(min_row=1, values_only=True)
    remitents = []
    count = 0
    for value in row_sheet:
        if value[2] != None: 
            count+=1
            remitents.append(value[2])
    
    for to in remitents:
        msg = message.Message()
        msg['Subject'] = 'Promocion - Codigo De Descuento!'
        msg['From'] = 'pocketvzla@gmail.com'
        msg.add_header('Content-Type', 'text/html')
        msg.set_payload(email_content)
        msg['To'] = to
        s.sendmail(msg['From'], [msg['To']], msg.as_string().encode('utf-8'))
        print('successfully sent email to {}:'.format(msg['To']))

    s.quit()

    print('Sended {} mails'.format(count))
